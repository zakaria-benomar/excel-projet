
from openpyxl import Workbook, load_workbook

import glob

r = 2
result = Workbook()
result_sheet = result.active

result_sheet.cell(row=1,column=1).value = 'Num Facture'
result_sheet.cell(row=1,column=2).value = 'Sté'
result_sheet.cell(row=1,column=3).value = 'Adresse'
result_sheet.cell(row=1,column=4).value = 'Date'
result_sheet.cell(row=1,column=5).value = 'Bon de Livraison'
result_sheet.cell(row=1,column=6).value = 'Bon de Commande'
result_sheet.cell(row=1,column=7).value = 'ID ARTICLE'
result_sheet.cell(row=1,column=8).value = 'Designation'
result_sheet.cell(row=1,column=9).value = 'Quantite'
result_sheet.cell(row=1,column=10).value = 'PU HT'
result_sheet.cell(row=1,column=11).value = 'Montant HT'
result_sheet.cell(row=1,column=12).value = 'Total HT'
result_sheet.cell(row=1,column=13).value = 'TVA'
result_sheet.cell(row=1,column=14).value = 'Montant TVA'
result_sheet.cell(row=1,column=15).value = 'Net a payer'


test_rows = 0
ligne = 0
column = 0
articles = []

path = "C:/Users/ATL/Downloads/FACTURE/FACTURE/"
chemain= ["FACTURE PAYER PAR ESPECE","FACTURE PAYER PAR CHEQ - EFFET","FACTURE NO PAYER"]

for folder in range(0, 3):

    if folder == 0:
        etat = " non payé"
    elif folder == 1:
        etat = " payé par chéque"
    else:
        etat = " payé par espece"

    for file in glob.glob(path + chemain[folder] + "/*.xlsx"):
        test_rows += 1
        book = load_workbook(file, data_only=True)
        sheet = book.active

        for row in range(1, 30):
            for col in range(1, 13):
                if str(sheet.cell(row=row, column=col).value).find('DESIGNATION') != -1:
                    ligne = row + 1
                    column = col
                    break


        def header():
            for row in range(1, 13):
                for col in range(1, 13):
                    data = str(sheet.cell(row=row, column=col).value)

                    if data.find('FACTURE N') != -1:
                        print(data)
                        result_sheet.cell(row=r, column=1).value = data




                    elif data.find('Sté') != -1:
                        data = data.split(":")
                        result_sheet.cell(row=r, column=2).value = data[1].strip()
                        result_sheet.cell(row=r, column=3).value = str(
                            sheet.cell(row=row + 1, column=col).value).strip()

                    elif data.find('CASABLANCA le') != -1:
                        data = data.split(":")
                        result_sheet.cell(row=r, column=4).value = data[1].strip()


                    elif data.find('BL N°') != -1:

                        result_sheet.cell(row=r, column=5).value = data.strip()

                    elif data.find('BC N°') != -1:
                        result_sheet.cell(row=r, column=6).value = data.strip() + etat

                    if result_sheet.cell(row=r, column=6).value == None:
                        result_sheet.cell(row=r, column=6).value = etat


        def footer():
            total_ht = 0
            montant_tva = 0
            net_a_payer = 0
            tva = 'TVA 20% VENTES'
            result_number = 1

            for row in range(1, 50):
                for col in range(1, 13):
                    data = str(sheet.cell(row=row, column=col).value)

                    if data.find('TOTAL,HT') != -1:
                        total_ht = sheet.cell(row=row + 1, column=col).value
                        for i in range(col + 1, col + 13):
                            if sheet.cell(row=row + 1, column=i).value != None:
                                result_number += 1
                                if result_number == 3:
                                    montant_tva = sheet.cell(row=row + 1, column=i).value
                                if result_number == 4:
                                    net_a_payer = sheet.cell(row=row + 1, column=i).value
                        break
            result_sheet.cell(row=r, column=12).value = total_ht
            result_sheet.cell(row=r, column=13).value = tva
            result_sheet.cell(row=r, column=14).value = montant_tva
            result_sheet.cell(row=r, column=15).value = net_a_payer


        while sheet.cell(row=ligne, column=column).value != None:
            coll = 8

            for col in range(column, 40):
                if sheet.cell(row=ligne, column=col).value != None:
                    result_sheet.cell(row=r, column=coll).value = sheet.cell(row=ligne, column=col).value
                    coll += 1

            header()
            footer()
            ligne += 1
            r += 1
            print(str(test_rows) +" "+ file)


result.save("C:/Users/ATL/Desktop/excel projet/result/all_factures_" + chemain[folder] + ".xlsx")




result_article = Workbook()
article_sheet = result_article.active

r_a = 2
id=1

article_sheet.cell(row=1,column=2).value = 'Article'
article_sheet.cell(row=1,column=3).value = 'Prix Unitaire'

final_result = load_workbook("C:/Users/ATL/Desktop/excel projet/result/all_factures_"+chemain[folder]+".xlsx")
sheet = final_result.active

for row in range(2,sheet.max_row+1):
    if str(sheet.cell(row=row , column=8).value) not in articles:
        articles.append(str(sheet.cell(row=row , column=8).value))
        article_sheet.cell(row=r_a, column=1).value =id
        article_sheet.cell(row=r_a ,column=2).value=str(sheet.cell(row=row , column=8).value)
        article_sheet.cell(row=r_a, column=3).value = sheet.cell(row=row, column=10).value
        for lgn in range(2,sheet.max_row+1):
            if str(article_sheet.cell(row=r_a ,column=2).value)== str(sheet.cell(row=lgn , column=8).value):
                sheet.cell(row=lgn, column=7).value = id
        r_a += 1
        id += 1

final_result.save('C:/Users/ATL/Desktop/excel projet/result/all_factures.xls')


result_article.save('C:/Users/ATL/Desktop/excel projet/result/Articles.xlsx')