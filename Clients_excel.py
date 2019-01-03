from openpyxl import Workbook, load_workbook


result = Workbook()
result_sheet = result.active

r = 2

result_sheet.cell(row=1,column=1).value = 'Sté'
result_sheet.cell(row=1,column=2).value = 'Adresse'

book = load_workbook('C:/Users/ATL/Desktop/excel projet/result/all_factures.xlsx')
sheet = book.active
client=[]
for row in range(2,sheet.max_row+1):
    if str(sheet.cell(row=row , column=2).value) not in client:
        client.append(str(sheet.cell(row=row , column=2).value))
        result_sheet.cell(row=r ,column=1).value=str(sheet.cell(row=row , column=2).value)
        result_sheet.cell(row=r, column=2).value = str(sheet.cell(row=row, column=3).value)
        r+=1


result.save('C:/Users/ATL/Desktop/excel projet/result/clients.xlsx')


